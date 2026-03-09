from openpyxl import load_workbook

def parse_student_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    students = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, email, roll, dept = row

        students.append({
            "name": name,
            "email": email,
            "roll": roll,
            "department": dept,
        })

    return students
