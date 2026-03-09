from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import load_workbook
from utils.serializer import serialize_mongo
import tempfile, os, shutil
from typing import Optional

from db.collections import faculty_collection, super_admins_collection
from schemas.faculty_schema import FacultyExcelRow
from core.guards import require_super_admin
from services.token_service import generate_reset_token
from services.email_service import send_email
from services.email_templates import build_password_email


router = APIRouter(
    prefix="/superadmin/faculty",
    tags=["Faculty"],
    dependencies=[Depends(require_super_admin)]
)

# ======================================================
# UPLOAD FACULTY (UNIVERSITY ISOLATED)
# ======================================================

@router.post("/upload")
async def upload_faculty(
    file: UploadFile = File(...),
    identity=Depends(require_super_admin)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")

    # 🔐 Resolve university from super admin (JWT sub)
    super_admin = await super_admins_collection.find_one(
        {"_id": ObjectId(identity["_id"])},
        {"university": 1}
    )

    if not super_admin or not super_admin.get("university"):
        raise HTTPException(403, "Super admin university not configured")

    university_id = super_admin["university"]["aishe_code"]
    university_name = super_admin["university"].get("name")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        temp_path = tmp.name

    wb = load_workbook(temp_path)
    ws = wb.active

    created, updated, skipped = 0, 0, 0
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            faculty = FacultyExcelRow(
                name=row[0],
                email=row[1],
                employee_id=row[2],
                department=row[3],
                designation=row[4],
                employment_type=row[5],
                joining_year=row[6],
                can_verify_certificates=row[7],
                can_verify_projects=row[8],
                can_verify_internships=row[9],
            )
        except Exception:
            skipped += 1
            errors.append(f"Invalid row: {row}")
            continue

        existing = await faculty_collection.find_one({
            "email": faculty.email,
            "academic.university_id": university_id
        })

        reset_token = generate_reset_token()
        expiry = datetime.utcnow() + timedelta(hours=24)

        doc = {
            "name": faculty.name,
            "email": faculty.email,
            "employee_id": faculty.employee_id,
            "role": "faculty",

            "password": None,
            "is_active": False,
            "created_at": datetime.utcnow(),
            "last_login": None,

            "academic": {
                "department": faculty.department,
                "designation": faculty.designation,
                "employment_type": faculty.employment_type,
                "joining_year": faculty.joining_year,
                "university_id": university_id,
                "university_name": university_name
            },

            "permissions": {
                "can_verify_certificates": faculty.can_verify_certificates,
                "can_verify_projects": faculty.can_verify_projects,
                "can_verify_internships": faculty.can_verify_internships
            },

            "verification_stats": {
                "verified_count": 0,
                "rejected_count": 0,
                "last_verified_at": None
            },

            "onboarding": {
                "reset_token": reset_token,
                "reset_token_exp": expiry,
                "email_verified": False,
                "onboarded_via": "excel_upload"
            },

            "meta": {
                "created_by": identity["_id"],  # ✅ FIXED
                "source": "bulk_upload",
                "last_updated": datetime.utcnow()
            }
        }

        if existing:
            await faculty_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": doc}
            )
            updated += 1
        else:
            await faculty_collection.insert_one(doc)
            created += 1

        await send_email(
            faculty.email,
            "Activate your Faculty Account",
            build_password_email(faculty.name, reset_token)
        )

    os.remove(temp_path)

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors
    }


# ======================================================
# LIST FACULTY (UNIVERSITY ISOLATED)
# ======================================================

@router.get("")
async def list_faculty(
    department: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    identity=Depends(require_super_admin)
):
    super_admin = await super_admins_collection.find_one(
        {"_id": ObjectId(identity["_id"])},
        {"university": 1}
    )

    if not super_admin or not super_admin.get("university"):
        raise HTTPException(403, "Super admin university not configured")
    print(super_admin)
    university_id = super_admin["university"]["aishe_code"]

    filters = {
        "academic.university_id": university_id
    }

    if department:
        filters["academic.department"] = department
    if is_active is not None:
        filters["is_active"] = is_active

    faculty = []
    async for doc in faculty_collection.find(
        filters,
        {"password": 0, "onboarding.reset_token": 0}
    ):
        doc["_id"] = str(doc["_id"])
        faculty.append(doc)
        

    return serialize_mongo(faculty)

@router.patch("/{faculty_id}/status")
async def update_faculty_status(
    faculty_id: str,
    is_active: bool = Query(...),
    identity=Depends(require_super_admin)
):
    
    aishe_code = identity["university"]["aishe_code"]

    result = await faculty_collection.update_one(
        {
            "_id": ObjectId(faculty_id),
            "academic.university_id": aishe_code  # 🔒 ISOLATION
        },
        {
            "$set": {
                "is_active": is_active,
                "meta.last_updated": datetime.utcnow()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(
            status_code=404,
            detail="Faculty not found or not part of your university"
        )

    return {
        "message": "Faculty status updated",
        "faculty_id": faculty_id,
        "is_active": is_active
    }

@router.post("/{faculty_id}/reset-password")
async def reset_faculty_password(
    faculty_id: str,
    identity=Depends(require_super_admin)
):
    aishe_code = identity["university"]["aishe_code"]

    faculty = await faculty_collection.find_one(
        {
            "_id": ObjectId(faculty_id),
            "academic.university_id": aishe_code  # 🔒 ISOLATION
        }
    )

    if not faculty:
        raise HTTPException(
            status_code=404,
            detail="Faculty not found or not part of your university"
        )

    reset_token = generate_reset_token()
    expiry = datetime.utcnow() + timedelta(hours=24)

    await faculty_collection.update_one(
        {"_id": faculty["_id"]},
        {
            "$set": {
                "onboarding.reset_token": reset_token,
                "onboarding.reset_token_exp": expiry,
                "is_active": False,
                "meta.last_updated": datetime.utcnow()
            }
        }
    )

    await send_email(
        to_email=faculty["email"],
        subject="Reset your Faculty Account Password",
        html=build_password_email(faculty["name"], reset_token)
    )

    return {
        "message": "Password reset email sent to faculty"
    }
