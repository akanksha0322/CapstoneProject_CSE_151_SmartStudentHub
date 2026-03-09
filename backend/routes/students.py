from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from bson import ObjectId
from datetime import datetime, timedelta
import tempfile, shutil, os, secrets

from core.guards import require_super_admin, require_student
from db.gridfs import fs
from db.collections import super_admins_collection, users_collection
from services.email_service import send_email
from core.config import Settings
from openpyxl import load_workbook

# =====================================================
# ROUTER
# =====================================================

router = APIRouter(
    prefix="/students",
    tags=["Students"]
)

FRONTEND_URL = Settings.frontend_URL


# =====================================================
# HELPERS
# =====================================================

def generate_reset_token():
    return secrets.token_urlsafe(32)


def build_password_email(name: str, token: str):
    link = f"{FRONTEND_URL}/set-password?token={token}"
    return f"""
    Hi {name},

    Your Smart Student Hub student account has been created.

    Please click the link below to set your password:

    {link}

    This link will expire in 24 hours.

    Regards,
    Smart Student Hub
    """


# =====================================================
# BULK STUDENT UPLOAD (SUPER ADMIN ONLY)
# =====================================================

@router.post(
    "/upload",
    dependencies=[Depends(require_super_admin)]
)
async def upload_students(
    file: UploadFile = File(...),
    admin: dict = Depends(require_super_admin)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")

    # 🔐 Resolve super admin (DB authoritative)
    admin_id = admin.get("user_id") or admin.get("_id")
    if not admin_id:
        raise HTTPException(401, "Invalid admin token")

    super_admin = await super_admins_collection.find_one(
        {"_id": ObjectId(admin_id), "role": "super_admin"},
        {"university.aishe_code": 1}
    )

    if not super_admin:
        raise HTTPException(401, "Super admin not found")

    aishe = super_admin.get("university", {}).get("aishe_code")
    if not aishe:
        raise HTTPException(400, "AISHE code not linked")

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            shutil.copyfileobj(file.file, tmp)
            temp_path = tmp.name

        wb = load_workbook(temp_path)
        ws = wb.active

        created, updated, skipped = 0, 0, 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                (
                    name,
                    email,
                    regno,
                    department,
                    program,
                    batch_year,
                    admission_year,
                    admission_type
                ) = row
            except Exception:
                skipped += 1
                errors.append(f"Invalid row format: {row}")
                continue

            if not all([name, email, regno, department, program, batch_year]):
                skipped += 1
                errors.append(f"Missing required fields: {row}")
                continue

            existing = await users_collection.find_one({
                "email": email,
                "academic.university_aishe": aishe
            })

            reset_token = generate_reset_token()
            expiry = datetime.utcnow() + timedelta(hours=24)

            student_doc = {
                "name": name,
                "email": email,
                "register_no": regno,
                "role": "student",
                "password": None,
                "is_active": False,
                "created_at": datetime.utcnow(),
                "last_login": None,

                "academic": {
                    "university_aishe": aishe,
                    "department": department,
                    "program": program,
                    "batch_year": batch_year,
                    "admission_year": admission_year or batch_year,
                    "admission_type": admission_type or "regular",
                    "current_year": None,
                    "semester": None
                },

                "profile": {
                    "phone": None,
                    "dob": None,
                    "gender": None,
                    "blood_group": None,
                    "address": None,
                    "photo_url": None
                },

                "status": {
                    "profile_completed": False,
                    "documents_uploaded": False,
                    "faculty_verified": False,
                    "graduated": False,
                    "blocked": False
                },

                "onboarding": {
                    "reset_token": reset_token,
                    "reset_token_exp": expiry,
                    "email_verified": False,
                    "onboarded_via": "excel_upload"
                },

                "activities": [],
                "certifications": [],
                "internships": [],
                "projects": [],
                "achievements": [],

                "meta": {
                    "created_by": admin_id,
                    "source": "bulk_upload",
                    "last_updated": datetime.utcnow()
                }
            }

            if existing:
                await users_collection.update_one(
                    {"_id": existing["_id"]},
                    {"$set": student_doc}
                )
                updated += 1
            else:
                await users_collection.insert_one(student_doc)
                created += 1

            await send_email(
                to_email=email,
                subject="Your Smart Student Hub Account",
                html=build_password_email(name, reset_token)
            )

        os.remove(temp_path)

        return {
            "message": "Upload completed",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10]
        }

    except Exception as e:
        raise HTTPException(500, f"Excel processing failed: {str(e)}")


# =====================================================
# DELETE CERTIFICATE (STUDENT ONLY)
# =====================================================

@router.delete(
    "/documents/certificate/{doc_id}",
    dependencies=[Depends(require_student)]
)
async def delete_certificate_document(
    doc_id: str,
    identity=Depends(require_student)
):
    student_id_raw = identity.get("_id") or identity.get("user_id")
    if not student_id_raw:
        raise HTTPException(401, "Invalid token")

    student_id = ObjectId(student_id_raw)

    try:
        cert_oid = ObjectId(doc_id)
    except Exception:
        raise HTTPException(400, "Invalid certificate id")

    student = await users_collection.find_one(
        {
            "_id": student_id,
            "role": "student",
            "certifications._id": cert_oid
        },
        {"certifications.$": 1}
    )

    if not student:
        raise HTTPException(404, "Certificate not found")

    cert = student["certifications"][0]

    # 🔒 Preserve your rule
    if cert.get("status") != "approved":
        raise HTTPException(403, "Only approved documents can be deleted")

    if cert.get("file_id"):
        try:
            fs.delete(ObjectId(cert["file_id"]))
        except Exception:
            pass

    await users_collection.update_one(
        {"_id": student_id},
        {"$pull": {"certifications": {"_id": cert_oid}}}
    )

    return {"message": "Certificate deleted successfully"}
